import datetime
import time
import os
from copy import copy
import pandas as pd
from methods.constants import EXCEL_COLUMNS
from methods.configs import RELATIVE_EXCELS_PATH
from methods.urls import reduce_url

def remove_bad_chars(df):
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    #illegal_chars = re.compile(r'[\x00-\x1F]')
    #return df.map(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    return df.map(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)

def categorize_postings_by_date(df):
    """
    Categorize postings by date ranges and return DataFrame sorted with separator markers.
    Categories: Last 7 days, Last 2 weeks, Last 4 weeks (2-4), Older than 4 weeks
    Adds a '_separator' column to mark group boundaries.
    """
    today = datetime.datetime.now()
    one_week_ago = today - datetime.timedelta(days=7)
    two_weeks_ago = today - datetime.timedelta(days=14)
    four_weeks_ago = today - datetime.timedelta(days=28)
    
    def get_category(row):
        try:
            if 'first_collected_on' in row and row['first_collected_on']:
                date = datetime.datetime.strptime(row['first_collected_on'], "%Y-%m-%d")
                if date >= one_week_ago:
                    return 0
                elif date >= two_weeks_ago:
                    return 1
                elif date >= four_weeks_ago:
                    return 2
                else:
                    return 3
            return 3
        except (ValueError, TypeError):
            return 3
    
    df['_category'] = df.apply(get_category, axis=1)
    df['_points'] = pd.to_numeric(df['points'], errors='coerce').fillna(0)
    df = df.sort_values(by=['_category', '_points'], ascending=[True, False])

    category_labels = {
        0: "Postings last 7 days - SCROLL FOR OLDER!!!",
        1: "Last 2 weeks postings:",
        2: "Last 4 weeks postings",
        3: "Older postings",
    }
    current_category = None
    separators = []
    
    for idx, (_, row) in enumerate(df.iterrows()):
        cat = row['_category']
        if cat != current_category:
            separators.append((idx, category_labels[cat]))
            current_category = cat
    
    df['_separator'] = ""
    for idx, label in separators:
        df.iloc[idx, df.columns.get_loc('_separator')] = label
    
    return df

def create_excel_report(df, path_excel = f"{RELATIVE_EXCELS_PATH}/", prefix ='postings',
                        added=None, keywords=None, throw_error=True, widths = None,
                        highlight_titles=None):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    excel_file_path = f"{path_excel}{prefix}_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    try:
        df_export = remove_bad_chars(df.drop(columns=['_category', '_points', '_separator'], errors='ignore'))
        df_export.to_excel(excel_file_path, index=False) #can also use engine='xlsxwriter'
        time.sleep(1)
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        if widths:
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width

        #Row bold, and columns bold
        for cell in sheet[1]:
            cell.font = cell.font + Font(bold=True)

        #Green text for highlighted postings (explicit list overrides "added")
        if highlight_titles is None:
            highlighted_titles = [instance["title"] for instance in added.values()] if added else []
        else:
            highlighted_titles = [title for title in highlight_titles if title]
        highlighted_titles = list(dict.fromkeys(highlighted_titles))
        separator_rows = {}
        
        for idx, row_num in enumerate(range(2, len(df) + 2)):
            row = sheet[row_num]
            if idx < len(df) and '_separator' in df.columns:
                separator_value = df.iloc[idx]['_separator']
                if separator_value:
                    separator_rows[row_num] = separator_value
            
            if row[0].value in highlighted_titles:
                for cell in row:
                    if cell.column_letter in ["A", "E"]:
                        font = copy(cell.font) #TODO pack this into a function - or class method e.g. cell.update_font(color = ..., ...)
                        font.color = "FF229F22"
                        cell.font = font
            if row[1].value and any(company_title in str(row[1].value).lower() for company_title in keywords.get("highlighted_company_titles", []) if isinstance(row[1].value, str)):
                for cell in row:
                    if cell.column_letter in ["B"]:
                        font = copy(cell.font)
                        font.color = "E30A0A"
                        cell.font = font

        #Bold columns (hyperlinks will be set later after separator insertion)
        for row in sheet.iter_rows(min_row=1, max_row=len(df)+1):
            for cell in row:
                if cell.column_letter in ["B", "D", "E"]:
                    cell.font = cell.font + Font(bold=True)

        #Style separator rows (insert actual separator rows)
        if separator_rows:
            separator_colors = {
                "Postings last 7 days - SCROLL FOR OLDER!!!": "CCE5FF", #light blue
                "Last 2 weeks postings:": "FFCCCC", #light yellow
                "Last 4 weeks postings": "FFE6CC", #light orange
                "Older postings": "E6F2E6" #light green
            }
            separator_font = Font(bold=True, size=22, color="000000")
            
            #Process separators from bottom to top to maintain row numbers
            for row_num in sorted(separator_rows.keys(), reverse=True):
                separator_label = separator_rows[row_num]
                #Insert a new row for the separator
                sheet.insert_rows(row_num)
                sep_row = sheet[row_num]
                
                #Set separator label in first cell and style entire row
                sep_row[1].value = separator_label
                sep_row[1].font = separator_font
                sep_row[1].alignment = Alignment(horizontal="center", vertical="center")
                
                color = separator_colors.get(separator_label, "CCCCCC")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                for cell in sep_row:
                    cell.fill = fill
                    cell.font = separator_font
                
                sheet.row_dimensions[row_num].height = 32

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == "F" and cell.value:
                    url = cell.value
                    if url and not (str(url).startswith("https://") or str(url).startswith("http://")):
                        url = "https://" + str(url)
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'
        workbook.save(excel_file_path)
    except Exception as e:
        if throw_error:
            raise e
        else:
            print(f"Could not create excel report at {excel_file_path}, error: {e}")
    return excel_file_path

def generate_outputs(results, added=None, keywords=None,
                     excel_cols=EXCEL_COLUMNS, excel_prefix="postings",
                     path_excel=RELATIVE_EXCELS_PATH, verbose=False,
                     highlight_first_collected_days=None):
    if not path_excel or results is None:
        return {"excel_file_path": None}

    df = pd.DataFrame.from_dict(results, orient='index')
    for col, vals in excel_cols.items():
        if col not in df.columns:
            df[col] = ""

    cols_ids = [vals['column'] for _, vals in excel_cols.items()]
    cols_ids_sorted = sorted(cols_ids, key=lambda x: (len(x), x))
    cols_sorted = [
        key for col_id in cols_ids_sorted
        for key, vals in excel_cols.items() if vals['column'] == col_id
    ]
    if 'first_collected_on' not in cols_sorted and 'first_collected_on' in df.columns:
        cols_sorted.append('first_collected_on')

    highlight_titles = None
    if highlight_first_collected_days is not None:
        cutoff = datetime.datetime.now() - datetime.timedelta(days=highlight_first_collected_days)
        highlight_titles = []
        for _, row in df.iterrows():
            date_str = row.get("first_collected_on")
            if not date_str:
                continue
            try:
                date_value = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            except (ValueError, TypeError):
                continue
            if date_value >= cutoff:
                title = row.get("title")
                if title:
                    highlight_titles.append(title)

    excel_df = df[cols_sorted].copy()
    excel_df = categorize_postings_by_date(excel_df)

    if 'url' in excel_df.columns:
        excel_df.loc[:, "url"] = excel_df["url"].apply(reduce_url)

    excel_df = excel_df.rename(columns={col: vals['as'] for col, vals in excel_cols.items()})
    excel_file_path = create_excel_report(
        excel_df,
        path_excel=path_excel, prefix=excel_prefix,
        added=added, keywords=keywords, throw_error=True,
        widths={vals['column']: vals.get('width', 20) for _, vals in excel_cols.items()},
        highlight_titles=highlight_titles,
    )
    if verbose:
        print(f"Excel report created at {excel_file_path}")

    return {"excel_file_path": excel_file_path}


def log_to_markdown(postings, log_file_path="newly_added_history.md"):
    """
    Prepend (add to top) postings (typically newly added ones) to a markdown log file.
    Newest entries appear at the top.
    
    Parameters:
    added (dict): Dictionary of newly added postings
    log_file_path (str): Path to the markdown log file
    """
    if not postings:
        return
    log_dir = os.path.dirname(log_file_path)
    if log_dir:
        os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    existing_content = ""
    if os.path.exists(log_file_path):
        with open(log_file_path, 'r', encoding='utf-8') as f:
            existing_content = f.read()
    with open(log_file_path, 'w', encoding='utf-8') as f:
        f.write(f"### {timestamp}\n")
        for key, posting in postings.items():
            if "title" in posting and "company" in posting:
                f.write(f"{posting['title']} - at - {posting['company']}\n")
        f.write("\n")
        if existing_content:
            f.write(existing_content)

def send_email(results, to_email, file_path = None, subject="Daily report", from_email=None):
    import smtplib
    #from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    new_postings = '\n'.join(result['title'] + '\n\t at ' + result['company'] for id, result in results['added'].items())
    companies_list = '\n'.join(results['companies'])
    
    body = "Dear User, here is your report of the current relevant job postings.\n\n" + \
            f"New posting not seen before:\n{new_postings}\n\nCurrently relevant companies:\n{companies_list}\n\n" + \
            "Best regards,\n\tHanicsBot"
    msg = MIMEMultipart()
    msg.attach(MIMEText(body, 'plain'))

    with open("config.txt", "r") as f:
        pw = f.readline().split("\n")[0]
        if not from_email:
            from_email = f.readline()

    msg['Subject'] = subject
    msg['From'] = from_email
    msg['To'] = to_email
    #msg['Bcc'] = from_email+"2"
    with open(file_path, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
        msg.attach(part)
    
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(from_email, pw)
        smtp.send_message(msg)