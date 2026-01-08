#Scrape website
import datetime
import pandas as pd
import time
from copy import copy
from typing import List, Type

import sys
import os
schedule_path = os.path.dirname(os.path.abspath(__file__))
parent_path = os.path.join(schedule_path, "..")
sys.path.append(parent_path)
from methods import files_io, postings_utils, datastruct_utils, sites, constants
from methods.constants import *

KEYWORDS = constants.BASE_PHRASES.copy()
RANKINGS = constants.BASE_KEYWORD_SCORING.copy()
SCRAPERS = [sites.KarriereAT, sites.Raiffeisen]
SALARY_BEARABLE = constants.SALARY_BEARABLE

def reduce_url(url):
    return url.split("www.")[1] if "www." in url else url.split("://")[1] if "://" in url else url

def categorize_postings_by_date(df):
    """
    Categorize postings by date ranges and return DataFrame sorted with separator markers.
    Categories: Last 2 weeks, Last 4 weeks (2-4), Older than 4 weeks
    Adds a '_separator' column to mark group boundaries.
    """
    today = datetime.datetime.now()
    two_weeks_ago = today - datetime.timedelta(days=14)
    four_weeks_ago = today - datetime.timedelta(days=28)
    
    def get_category(row):
        try:
            if 'first_collected_on' in row and row['first_collected_on']:
                date = datetime.datetime.strptime(row['first_collected_on'], "%Y-%m-%d")
                if date >= two_weeks_ago:
                    return 0
                elif date >= four_weeks_ago:
                    return 1
                else:
                    return 2
            return 2
        except (ValueError, TypeError):
            return 2
    
    df['_category'] = df.apply(get_category, axis=1)
    df['_points'] = pd.to_numeric(df['points'], errors='coerce').fillna(0)
    df = df.sort_values(by=['_category', '_points'], ascending=[True, False])

    category_labels = {0: "Last 2 weeks postings", 1: "Last 4 weeks postings", 2: "Older postings"}
    current_category = None
    separators = []
    
    for idx, (i, row) in enumerate(df.iterrows()):
        cat = row['_category']
        if cat != current_category:
            separators.append((idx, category_labels[cat]))
            current_category = cat
    
    df['_separator'] = ""
    for idx, label in separators:
        df.iloc[idx, df.columns.get_loc('_separator')] = label
    
    return df

def reorder_locations(locations, primary = LOCATIONS_DESIRED):
    if not isinstance(locations, list) or not locations:
        return locations
    first_locs = []
    other_locs = []
    for loc in locations:
        if isinstance(loc, str) and any(city in loc.lower() for city in primary):
            first_locs.append(loc)
        else:
            other_locs.append(loc)
    return first_locs + other_locs

def remove_bad_chars(df):
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    #illegal_chars = re.compile(r'[\x00-\x1F]')
    return df.map(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)

def replace_chars(df):
    return df.map(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

def create_excel_report(df, path_excel = f"{RELATIVE_EXCELS_PATH}/", prefix ='postings',
                        added=None, keywords=None, throw_error=True, widths = None):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    excel_file_path = f"{path_excel}{prefix}_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    try:
        df_export = remove_bad_chars(df.drop(columns=['_category', '_points', '_separator'], errors='ignore'))
        df_export.to_excel(excel_file_path, index=False) #can also use engine='xlsxwriter'
        time.sleep(1)
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        if widths:
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width

        #Row bold, and columns bold
        for cell in sheet[1]:
            cell.font = cell.font + Font(bold=True)

        #Make those rows green text which are in the added postings, and highlight title companies with red names
        added_titles = [instance["title"] for instance in added.values()] if added else []
        separator_rows = {}
        
        for idx, row_num in enumerate(range(2, len(df) + 2)):
            row = sheet[row_num]
            if idx < len(df) and '_separator' in df.columns:
                separator_value = df.iloc[idx]['_separator']
                if separator_value:
                    separator_rows[row_num] = separator_value
            
            if row[0].value in added_titles:
                for cell in row:
                    if cell.column_letter in ["A", "E"]:
                        font = copy(cell.font) #TODO pack this into a function - or class method e.g. cell.update_font(color = ..., ...)
                        font.color = "FF229F22"
                        cell.font = font
            if row[1].value and any(company_title in str(row[1].value).lower() for company_title in keywords.get("highlighted_company_titles", []) if isinstance(row[1].value, str)):
                for cell in row:
                    if cell.column_letter in ["B"]:
                        font = copy(cell.font)
                        font.color = "E30A0A"
                        cell.font = font

        #Bold columns (hyperlinks will be set later after separator insertion)
        for row in sheet.iter_rows(min_row=1, max_row=len(df)+1):
            for cell in row:
                if cell.column_letter in ["B", "D", "E"]:
                    cell.font = cell.font + Font(bold=True)

        #Style separator rows (insert actual separator rows)
        if separator_rows:
            separator_colors = {
                "Last 2 weeks postings": "FFCCCC", #light yellow
                "Last 4 weeks postings": "FFE6CC", #light orange
                "Older postings": "E6F2E6" #light green
            }
            separator_font = Font(bold=True, size=22, color="000000")
            
            #Process separators from bottom to top to maintain row numbers
            for row_num in sorted(separator_rows.keys(), reverse=True):
                separator_label = separator_rows[row_num]
                #Insert a new row for the separator
                sheet.insert_rows(row_num)
                sep_row = sheet[row_num]
                
                #Set separator label in first cell and style entire row
                sep_row[1].value = separator_label
                sep_row[1].font = separator_font
                sep_row[1].alignment = Alignment(horizontal="center", vertical="center")
                
                color = separator_colors.get(separator_label, "CCCCCC")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                for cell in sep_row:
                    cell.fill = fill
                    cell.font = separator_font
                
                sheet.row_dimensions[row_num].height = 32

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == "F" and cell.value:
                    url = cell.value
                    if url and not (str(url).startswith("https://") or str(url).startswith("http://")):
                        url = "https://" + str(url)
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'
        workbook.save(excel_file_path)
    except Exception as e:
        if throw_error:
            raise e
        else:
            print(f"Could not create excel report at {excel_file_path}, error: {e}")
    return excel_file_path

def scrape_sites(scrapers: List[Type[sites.BaseScraper]] = SCRAPERS, keywords=KEYWORDS, rankings=RANKINGS, salary_bearable=SALARY_BEARABLE,
                 verbose=False, verbose_data_gathering=False, **kwargs):
    results = {}
    for scraper_class in scrapers:
        scraper = scraper_class(keywords=keywords, rankings=rankings, salary_bearable=salary_bearable,
                                extra_keywords=kwargs.get(f"{scraper_class.__name__.lower()}_extra_keywords", kwargs.get("extra_keywords", {})),
                                extra_titlewords=kwargs.get(f"{scraper_class.__name__.lower()}_extra_titlewords", kwargs.get("extra_titlewords", [])),
                                extra_locations=kwargs.get(f"{scraper_class.__name__.lower()}_extra_locations", kwargs.get("extra_locations", [])),
                                locations_secondary=kwargs.get(f"{scraper_class.__name__.lower()}_locations_secondary", kwargs.get("locations_secondary", LOCATIONS_SECONDARY)),
                                rules=kwargs.get(f"{scraper_class.__name__.lower()}_rules", kwargs.get("rules", BASE_RULES)),
                                transformations = kwargs.get(f"{scraper_class.__name__.lower()}_transformations", kwargs.get("transformations", [])),
                                )
        run_results = scraper.gather_data(descriptions=True, verbose=verbose_data_gathering)
        if verbose:
            print(f"Found {len(run_results)} postings on {scraper_class.__name__}")
        results = {**results, **run_results}
    return results

def get_postings(scrapers = SCRAPERS, keywords =KEYWORDS, rankings=RANKINGS, salary_bearable=SALARY_BEARABLE, path=f"{RELATIVE_POSTINGS_PATH}/",
                 path_excel=f"{RELATIVE_EXCELS_PATH}/", excel_prefix ="postings", verbose=False, verbose_data_gathering=False, threshold = 0.01,
                 col_order = COLUMN_ORDER, excel_cols = EXCEL_COLUMNS, efficient_storing = True, only_thresholded_stored = False,
                 current_postings_filename = CURRENT_POSTINGS_FILENAME, newly_added_postings_filename = NEWLY_ADDED_POSTINGS_FILENAME,
                 postings_history_filename = POSTINGS_HISTORY_FILENAME, **kwargs):
    keywords['titlewords'] = list(set(keywords['titlewords']))
    results = scrape_sites(scrapers = scrapers, keywords=keywords, rankings=rankings, salary_bearable=salary_bearable,
                            verbose=verbose, verbose_data_gathering=verbose_data_gathering, **kwargs)
    results = datastruct_utils.reorder_dict(results, col_order, nested=True)
    results = {k: v for k, v in sorted(results.items(), key=lambda item: item[1]["points"], reverse=True)}
    if verbose:
        print(f"Found {len(results)} results in total")
    
    previous_file_name = "postings_history.json"
    if not os.path.exists(f"{path}{previous_file_name}"):
        previous_file_name = files_io.get_filename_from_dir(path, index = -1) #if there is no file, returns None
    added, removed = postings_utils.get_added_and_removed(results, f'{path}{previous_file_name}' if previous_file_name else [])

    if only_thresholded_stored:
        added = postings_utils.threshold_postings_by_points(added, points_threshold=threshold)
        removed = postings_utils.threshold_postings_by_points(removed, points_threshold=threshold)
        results = postings_utils.threshold_postings_by_points(results, points_threshold=threshold)

    files_io.save_data(results, path=path, name=current_postings_filename, with_timestamp=False, )

    if not efficient_storing or not previous_file_name:
        files_io.save_data(results, name=postings_history_filename if efficient_storing else "",
                           path=path, with_timestamp=False if efficient_storing else True)
    else: #efficient storing
        history = postings_utils.unify_postings(folder_path=path) #includes newly_added_postings
        for key in history.keys():
            dt_last = datetime.datetime.strptime(history[key]['last_collected_on'], "%Y-%m-%d")
            dt_first = datetime.datetime.strptime(history[key]['first_collected_on'], "%Y-%m-%d")
            history[key]['timespan_days'] = (dt_last - dt_first).days
        for key in results.keys() & history.keys():
            results[key]['first_collected_on'] = history[key]['first_collected_on']
            results[key]['last_collected_on'] = history[key]['last_collected_on']
            results[key]['timespan_days'] = history[key]['timespan_days']
        for key in added.keys() & history.keys():
            added[key]['first_collected_on'] = history[key]['first_collected_on']
            added[key]['last_collected_on'] = history[key]['last_collected_on']
            added[key]['timespan_days'] = history[key]['timespan_days']
        files_io.save_data(history, path=path, name=postings_history_filename, with_timestamp=False)
        #Re-run: adding first and last collected on to results
        files_io.save_data(results, path=path, name=current_postings_filename, with_timestamp=False)

    files_io.save_data(added, path=path, name=newly_added_postings_filename, with_timestamp=False, )

    if threshold is not None:
        added = postings_utils.threshold_postings_by_points(added, points_threshold=threshold)
        removed = postings_utils.threshold_postings_by_points(removed, points_threshold=threshold)
        results = postings_utils.threshold_postings_by_points(results, points_threshold=threshold)
    if verbose:
        print(f"Found {len(results)} postings. Added {len(added)} postings, removed {len(removed)} postings." + " above threshold" if threshold is not None else "")

    df = pd.DataFrame.from_dict(results, orient='index')
    df["locations"] = df["locations"].apply(lambda x: x if isinstance(x, list) else [])
    df["locations"] = df["locations"].apply(reorder_locations, primary=kwargs.get("locations_primary", LOCATIONS_DESIRED))
    companies = list(df[df['locations'].apply(lambda locs: any("wien" in loc.lower() or "vienna" in loc.lower() for loc in locs))]['company'].unique())

    excel_file_path = None
    if path_excel:
        #TODO put also into a function
        for col, vals in excel_cols.items():
            if not col in df.columns :
                df[col] = ""
        cols_ids = [vals['column'] for key, vals in excel_cols.items()]
        cols_ids_sorted = sorted(cols_ids, key=lambda x: (len(x), x))
        cols_sorted = [key for col_id in cols_ids_sorted for key, vals in excel_cols.items() if vals['column'] == col_id]
        if 'first_collected_on' not in cols_sorted and 'first_collected_on' in df.columns:
            cols_sorted.append('first_collected_on')
        excel_df = df[cols_sorted].copy()
        excel_df = categorize_postings_by_date(excel_df)

        if 'url' in excel_df.columns:
            excel_df.loc[:, "url"] = excel_df["url"].apply(lambda x: reduce_url(x))
        
        excel_df = excel_df.rename(columns={col: vals['as'] for col, vals in excel_cols.items()})
        excel_file_path = create_excel_report(excel_df, path_excel = path_excel, prefix = excel_prefix,
                                              added = added, keywords = keywords, throw_error=True,
                                              widths = {vals['column']: vals.get('width', 20)
                                                        for col, vals in excel_cols.items()})
        if verbose:
            print(f"Excel report created at {excel_file_path}")

    output_dict = {
        "results": results,
        "added": added, #above threshold
        "removed": removed, #above threshold
        "companies": companies,
        "excel_file_path": excel_file_path,
    }
    return output_dict

def send_email(results, to_email, file_path = None, subject="Daily report", from_email=None):
    import smtplib
    #from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    new_postings = '\n'.join(result['title'] + '\n\t at ' + result['company'] for id, result in results['added'].items())
    companies_list = '\n'.join(results['companies'])
    
    body = "Dear User, here is your report of the current relevant job postings.\n\n" + \
            f"New posting not seen before:\n{new_postings}\n\nCurrently relevant companies:\n{companies_list}\n\n" + \
            "Best regards,\n\tHanicsBot"
    msg = MIMEMultipart()
    msg.attach(MIMEText(body, 'plain'))

    with open("config.txt", "r") as f:
        pw = f.readline().split("\n")[0]
        if not from_email:
            from_email = f.readline()

    msg['Subject'] = subject
    msg['From'] = from_email
    msg['To'] = to_email
    #msg['Bcc'] = from_email+"2"
    with open(file_path, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
        msg.attach(part)
    
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(from_email, pw)
        smtp.send_message(msg)

if __name__ == "__main__":
    data = get_postings()
    print("Found:" , len(data["results"]), "postings")
    #print(data) #find a nice format
