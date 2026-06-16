"""Excel-specific output helpers."""

import datetime
import time
from copy import copy

import pandas as pd

from jobscraping.config.configs import RELATIVE_EXCELS_PATH


def remove_bad_chars(df: pd.DataFrame) -> pd.DataFrame:
    """Strip illegal characters before exporting to Excel."""
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # pylint: disable=import-outside-toplevel

    return df.map(
        lambda x: ILLEGAL_CHARACTERS_RE.sub(r"", x) if isinstance(x, str) else x
    )


def create_excel_report(
    df: pd.DataFrame,
    path_excel: str = f"{RELATIVE_EXCELS_PATH}/",
    prefix: str = "postings",
    added: dict | None = None,
    keywords: dict | None = None,
    throw_error: bool = True,
    widths: dict | None = None,
    highlight_titles: list | None = None,
) -> str:
    """Create and style an Excel report from a DataFrame."""
    from openpyxl import load_workbook  # pylint: disable=import-outside-toplevel
    from openpyxl.styles import Alignment, Font, PatternFill  # pylint: disable=import-outside-toplevel

    excel_file_path = f"{path_excel}{prefix}_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    try:
        df_export = remove_bad_chars(
            df.drop(columns=["_category", "_points", "_separator"], errors="ignore")
        )
        df_export.to_excel(excel_file_path, index=False)
        time.sleep(1)
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        if widths:
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width

        for cell in sheet[1]:
            cell.font = cell.font + Font(bold=True)

        if highlight_titles is None:
            highlighted_titles = [instance["title"] for instance in added.values()] if added else []
        else:
            highlighted_titles = [title for title in highlight_titles if title]
        highlighted_titles = list(dict.fromkeys(highlighted_titles))
        separator_rows = {}

        keywords = keywords or {}
        for idx, row_num in enumerate(range(2, len(df) + 2)):
            row = sheet[row_num]
            if idx < len(df) and "_separator" in df.columns:
                separator_value = df.iloc[idx]["_separator"]
                if separator_value:
                    separator_rows[row_num] = separator_value

            if row[0].value in highlighted_titles:
                for cell in row:
                    if cell.column_letter in ["A", "E"]:
                        font = copy(cell.font)
                        font.color = "FF229F22"
                        cell.font = font
            if row[1].value and any(
                company_title in str(row[1].value).lower()
                for company_title in keywords.get("highlighted_company_titles", [])
                if isinstance(row[1].value, str)
            ):
                for cell in row:
                    if cell.column_letter in ["B"]:
                        font = copy(cell.font)
                        font.color = "E30A0A"
                        cell.font = font

        for row in sheet.iter_rows(min_row=1, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter in ["B", "D", "E"]:
                    cell.font = cell.font + Font(bold=True)

        if separator_rows:
            separator_colors = {
                "Postings last 7 days - SCROLL FOR OLDER!!!": "CCE5FF",
                "Last 2 weeks postings:": "FFCCCC",
                "Last 4 weeks postings": "FFE6CC",
                "Older postings": "E6F2E6",
            }
            separator_font = Font(bold=True, size=22, color="000000")

            for row_num in sorted(separator_rows.keys(), reverse=True):
                separator_label = separator_rows[row_num]
                sheet.insert_rows(row_num)
                sep_row = sheet[row_num]

                sep_row[1].value = separator_label
                sep_row[1].font = separator_font
                sep_row[1].alignment = Alignment(horizontal="center", vertical="center")

                color = separator_colors.get(separator_label, "CCCCCC")
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                for cell in sep_row:
                    cell.fill = fill
                    cell.font = separator_font

                sheet.row_dimensions[row_num].height = 32

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == "F" and cell.value:
                    url = cell.value
                    if url and not (
                        str(url).startswith("https://") or str(url).startswith("http://")
                    ):
                        url = "https://" + str(url)
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
        workbook.save(excel_file_path)
    except Exception as exc:
        if throw_error:
            raise
        print(f"Could not create excel report at {excel_file_path}, error: {exc}")
    return excel_file_path
